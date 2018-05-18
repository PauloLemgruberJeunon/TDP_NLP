import projectFiles.Utils.textPreProcessing as txtPreProc
from projectFiles import utils

import xlsxwriter
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string

import math
import numpy as np


class MyExcelFileWrite:

    def __init__(self, workbook_path, workbook_name):
        self.my_path = workbook_path
        self.my_name = workbook_name

        self.sheet_dict = {}

        self.workbook = xlsxwriter.Workbook(self.my_path + self.my_name)

    def __enter__(self):
        return self

    def __exit__(self):
        self.close_workbook()

    def add_new_worksheet(self, name):
        #  Add a new worksheet to a "workbook" and return it (returns the worksheet)
        if name not in self.sheet_dict:
            self.sheet_dict[name] = self.workbook.add_worksheet(name)
        else:
            print('[WARNING] Sheet with this name already added')

    def close_workbook(self):
        #  Closes the workbook (mandatory by the end of its use)
        self.workbook.close()

    def write_matrix_in_xlsx(self, worksheet_name, matrix, row_index_name_dict, column_index_name_dict):
        """
        This function will print the matrix in a excel spreadsheet
        :param row_index_name_dict: The dictionary that stores the name of the rows (nouns) to print them in the sheet
        :param column_index_name_dict: The dictionary that stores the name of the columns (verbs)
         to print them in the sheet
        :param matrix: The co-occurrence matrix that will be written to the arquive
        :param worksheet_name: The worksheet that the matrix will be written to
        :return: Nothing
        """
        if worksheet_name in self.sheet_dict:
            worksheet = self.sheet_dict[worksheet_name]
        else:
            print('[WARNING] Worksheet was not found .... doing nothing')
            return

        row_len = len(row_index_name_dict)
        column_len = len(column_index_name_dict)

        for i in range(row_len):
            worksheet.write(i+1, 0, row_index_name_dict[i])

        for i in range(column_len):
            worksheet.write(0, i+1, column_index_name_dict[i])

        for i in range(row_len):
            for j in range(column_len):
                worksheet.write(i + 1, j + 1, matrix[i][j])

    def write_dict_matrix_in_xlsx(self, worksheet_name, dict_matrix, row_key_list, column_key_list):

        if worksheet_name in self.sheet_dict:
            worksheet = self.sheet_dict[worksheet_name]
        else:
            print('[WARNING] Worksheet was not found .... doing nothing')
            return

        i = 1
        for key in row_key_list:
            worksheet.write(i, 0, key)
            i += 1

        j = 1
        for key in column_key_list:
            worksheet.write(0, j, key)
            j += 1

        i = 1
        for key in row_key_list:
            j = 1
            for key2 in column_key_list:
                worksheet.write(i, j, dict_matrix[key][key2])
                j += 1
            i += 1

    def write_verb_filtered_arrays(self, worksheet_name, verb_filtered_arrays, nouns_from_verb_arrays, ordered_verbs):

        workbook = self.workbook

        if worksheet_name in self.sheet_dict:
            worksheet = self.sheet_dict[worksheet_name]
        else:
            print('[WARNING] Worksheet was not found .... doing nothing')
            return

        xlsx_column = 0
        div_format = workbook.add_format({'bg_color': 'green'})
        for verb in ordered_verbs:
            nouns = nouns_from_verb_arrays[verb]
            nouns_lenght = len(nouns)

            xlsx_row = 1
            for noun in nouns:
                worksheet.write(xlsx_row, xlsx_column, noun)
                xlsx_row += 1

            xlsx_row = 0
            xlsx_column += 1
            worksheet.write(xlsx_row, xlsx_column, verb)
            xlsx_row = 1

            values = verb_filtered_arrays[verb]
            for value in values:
                worksheet.write(xlsx_row, xlsx_column, value)
                xlsx_row += 1

            xlsx_row = 0
            xlsx_column += 2
            for i in range(nouns_lenght):
                worksheet.write_blank(xlsx_row, xlsx_column, '', div_format)
                xlsx_row += 1

            xlsx_column += 2


def save_matrix_in_xlsx(cooc_matrix, pure_matrix, path_to_output_xlsx, workbook_name):
    # The worksheet will be saved in a excel workbook with the file name and location equal to the string below
    with MyExcelFileWrite(path_to_output_xlsx, workbook_name) as workbook:
        workbook.add_new_worksheet('pure_frequency_count')

        # Just a test worksheet with the pure frequency numbers of the pairs of verbs and nouns
        workbook.write_matrix_in_xlsx('pure_frequency_count', pure_matrix, utils.invert_dictionary(cooc_matrix.noun_rows),
                                      utils.invert_dictionary(cooc_matrix.verb_columns))

        # In the lines below the worksheets will be created and associated to the workbook
        workbook.add_new_worksheet('cooc_matrix_full')
        workbook.add_new_worksheet('verb_filtered_arrays')

        # worksheet2 = get_new_worksheet('soc_pmi_matrix', workbook)

        inverted_matrix_noun_rows = utils.invert_dictionary(cooc_matrix.noun_rows)

        workbook.write_matrix_in_xlsx('cooc_matrix_full', cooc_matrix.matrix, inverted_matrix_noun_rows,
                                      utils.invert_dictionary(cooc_matrix.verb_columns))

        ordered_verbs = utils.sort_dict(cooc_matrix.verb_filtered_arrays)
        ordered_verbs = [ordered_verbs[l][0] for l in range(len(ordered_verbs))]
        workbook.write_verb_filtered_arrays('verb_filtered_arrays', cooc_matrix.verb_filtered_arrays,
                                            cooc_matrix.nouns_from_verb_arrays, ordered_verbs)


def find_associated_verbs_in_xlsx_sheet(full_noun_and_verb_list, sheet):
    temp_noun_list = []
    temp_verb_list = []

    for i in range(1, 13):
        temp_noun_list += sheet['Nouns' + str(i)].values.tolist().copy()
        temp_verb_list += sheet['Verbs' + str(i)].values.tolist().copy()

    temp_noun_list = [str(temp_noun_list[i]).strip() for i in range(len(temp_noun_list))]
    temp_verb_list = [str(temp_verb_list[i]).strip() for i in range(len(temp_verb_list))]

    for j in range(len(full_noun_and_verb_list)):
        for k in range(len(temp_noun_list)):
            if full_noun_and_verb_list[j] == temp_noun_list[k]:
                full_noun_and_verb_list[j] = (full_noun_and_verb_list[j], '[' + temp_verb_list[k] + ']')
                break


def find_nature_of_entities(full_noun_list, wb):
    nature_of_entities_list = []
    list_of_entities_lists = []
    list_of_nature_of_entities_list = []

    for i in range(1, 7):
        curr_sheet = wb.parse("stage " + str(i))
        list_of_entities_lists.append(curr_sheet['Entities'].values.tolist().copy())
        list_of_nature_of_entities_list.append(curr_sheet['Nature of Entities (Basic)'].values.tolist().copy())

    for entity in full_noun_list:

        for i in range(6):
            should_break = False

            for j in range(len(list_of_entities_lists[i])):
                if isinstance(list_of_nature_of_entities_list[i][j], float):
                    break

                if entity.strip() == list_of_entities_lists[i][j].strip():
                    nature_of_entities_list.append(list_of_nature_of_entities_list[i][j].lower())
                    should_break = True
                    break

            if should_break:
                break

    return nature_of_entities_list


def read_all_stages(input_address, nouns_for_hypernyms_file_name, list_of_verbs_file_name):
    content_dict = {}

    wb = pd.ExcelFile(input_address + nouns_for_hypernyms_file_name)

    wb2 = pd.ExcelFile(input_address + list_of_verbs_file_name)
    sheet2 = wb2.parse('sheet2')

    for i in range(1, 7):
        curr_sheet = wb.parse("stage " + str(i))
        noun_list = curr_sheet['Active nouns'].values.tolist().copy()
        department_list = curr_sheet['Department'].values.tolist().copy()
        full_noun_and_verb_list = curr_sheet['Entities'].values.tolist().copy()
        synset_list = curr_sheet['SUMO word ID'].values.tolist().copy()
        final_noun_list = curr_sheet['Curated nouns'].values.tolist().copy()
        nature_of_entities_list = curr_sheet['Nature of Entities (Basic)'].values.tolist().copy()

        for l in range(len(noun_list)):
            if isinstance(final_noun_list[l], str):
                noun_list[l] = final_noun_list[l]

        noun_list = txtPreProc.strip_list([txtPreProc.lemmatize_word(noun) for noun in noun_list])
        department_list = txtPreProc.strip_list(department_list)
        full_noun_and_verb_list = txtPreProc.strip_list(full_noun_and_verb_list)
        nature_of_entities_list = txtPreProc.strip_list(nature_of_entities_list)

        j = 0
        loop_size = len(noun_list)
        while j < loop_size:
            if isinstance(synset_list[j], float):
                if math.isnan(synset_list[j]):
                    synset_list[j] = 0.0

            if isinstance(noun_list[j], str) is False or noun_list[j] == 'nan':
                del noun_list[j:]
                del department_list[j:]
                del full_noun_and_verb_list[j:]
                del synset_list[j:]
                del nature_of_entities_list[j:]
                break
            j += 1

        synset_list = [int(k) for k in synset_list]

        find_associated_verbs_in_xlsx_sheet(full_noun_and_verb_list, sheet2)

        content_dict['stage' + str(i)] = {'noun_list': noun_list, 'department_list': department_list,
                                          'synset_list': synset_list,
                                          'full_noun_and_verb_list': full_noun_and_verb_list,
                                          'nature_of_entities_list': nature_of_entities_list}

    return content_dict


def read_all_nouns(input_address, nouns_for_hypernyms_file_name, list_of_verbs_file_name):
    wb = pd.ExcelFile(input_address + nouns_for_hypernyms_file_name)
    sheet = wb.parse("all nouns")

    noun_list = sheet['Active nouns'].values.tolist().copy()
    department_list = sheet['Department'].values.tolist().copy()
    full_noun_and_verb_list = sheet['Entities'].values.tolist().copy()
    synset_list = sheet['SUMO word ID'].values.tolist().copy()
    final_noun_list = sheet['Final noun'].values.tolist().copy()

    nature_of_entities_list = find_nature_of_entities(full_noun_and_verb_list, wb)

    noun_list = txtPreProc.strip_list(string_list=noun_list)
    department_list = txtPreProc.strip_list(string_list=department_list)
    full_noun_and_verb_list = txtPreProc.strip_list(string_list=full_noun_and_verb_list)
    nature_of_entities_list = txtPreProc.strip_list(string_list=nature_of_entities_list)

    wb2 = pd.ExcelFile(input_address + list_of_verbs_file_name)
    sheet2 = wb2.parse('sheet2')

    find_associated_verbs_in_xlsx_sheet(full_noun_and_verb_list, sheet2)

    for i in range(len(noun_list)):
        if isinstance(final_noun_list[i], str):
            noun_list[i] = final_noun_list[i]

        noun_list[i] = noun_list[i].lower()
        noun_list[i] = txtPreProc.lemmatize_word(noun_list[i])

    i = 0
    while_size = len(noun_list)
    while i < while_size:
        noun_list[i] = noun_list[i].replace(" ", "")

        if isinstance(synset_list[i], float):
            if math.isnan(synset_list[i]):
                synset_list[i] = 0.0

        i += 1

    synset_list = [int(k) for k in synset_list]

    content_dict = {"full_noun_and_verb_list": full_noun_and_verb_list, "noun_list": noun_list,
                    "department_list": department_list, "synset_list": synset_list,
                    'nature_of_entities_list': nature_of_entities_list}

    return content_dict


def read_verb_frequency_from_hypernym_graph(file_path, file_name):
    wb = load_workbook(filename=file_path + file_name, read_only=True)
    sheet = wb['Associated VerbNounPairs']

    rows = sheet.rows
    content_dict = {}
    row_list = list(rows)
    i = 0
    while i < len(row_list):

        curr_cell_list = list(row_list[i])
        if curr_cell_list[0].value is None:
            break

        curr_key = curr_cell_list[1].value
        content_dict[curr_key] = {}

        content_dict[curr_key]['hypernymVerb'] = curr_cell_list[2].value

        curr_cell_list = list(row_list[i + 1])
        content_dict[curr_key]['level'] = curr_cell_list[1].value

        curr_cell_list = list(row_list[i + 2])
        content_dict[curr_key]['pairSize'] = int(curr_cell_list[1].value)

        verb_cell_list = list(row_list[i + 3])
        noun_cell_list = list(row_list[i + 4])
        content_dict[curr_key]['verbList'] = []
        content_dict[curr_key]['nounList'] = []
        j = 1
        while True:
            if verb_cell_list[j].value is None:
                break
            elif verb_cell_list[j].value == '-':
                j += 1
                continue
            content_dict[curr_key]['verbList'].append(verb_cell_list[j].value.strip(']').strip('['))
            content_dict[curr_key]['nounList'].append(noun_cell_list[j].value)

            j += 1

        curr_cell_list = list(row_list[i + 5])
        content_dict[curr_key]['fatherList'] = []
        j = 1
        while True:
            if curr_cell_list[j].value is None:
                break

            content_dict[curr_key]['fatherList'].append(curr_cell_list[j].value)
            j += 1

        i += 9

    return content_dict


def get_column_names(rows, verb_columns):
    counter = 0
    for row in rows:
        for cell in row:
            value = cell.value
            if value is not None:
                verb_columns[value] = counter
                counter += 1
        break


def complete_the_loading(rows, noun_rows, matrix):
    counter = 0
    i = 0
    for row in rows:
        row_jump = True
        j = 0

        for cell in row:
            value = cell.value
            if row_jump:
                noun_rows[value] = counter
                counter += 1
                row_jump = False
            else:
                matrix[i][j] = value
                j += 1

        i += 1


def load_from_wb(workbook_name):
    """
    Function used to load the matrix data from a xlsx archive. It is faster then generating it from text
    :param workbook_name: path and name of the xlsx file to use to extract information
    :return: Returns a dictionary containing the information to create a 'CoocMatrix' object
    """
    # Load workbook
    wb = load_workbook(filename=workbook_name, read_only=True)
    print('wb loaded')

    # Load worksheets
    ws = wb['cooc_matrix_full']
    ws_soc_pmi = wb['soc_pmi_matrix']

    # Calculate the matrix dimensions and transform the excel's coordinates to matrix coordinates (to only integers)
    rows = ws.rows
    matrix_dim = ws.calculate_dimension().split(':')
    rows_count = coordinate_from_string(matrix_dim[1])[1] - 1
    column_count = column_index_from_string(coordinate_from_string(matrix_dim[1])[0]) - 1

    matrix = np.empty((rows_count, column_count))

    print('matrix allocated')

    noun_rows = {}
    verb_columns = {}

    get_column_names(rows, verb_columns)
    print('get_column_names completed')
    complete_the_loading(rows, noun_rows, matrix)
    print('complete_the_loading completed')

    rows = ws_soc_pmi.rows
    matrix_dim = ws_soc_pmi.calculate_dimension().split(':')
    rows_count = coordinate_from_string(matrix_dim[1])[1] - 1
    column_count = column_index_from_string(coordinate_from_string(matrix_dim[1])[0]) - 1

    soc_pmi_matrix = np.empty((rows_count, column_count))

    soc_pmi_noun_rows = {}
    soc_pmi_verb_columns = {}

    get_column_names(rows, soc_pmi_verb_columns)
    print('get_column_names completed 2')
    complete_the_loading(rows, soc_pmi_noun_rows, soc_pmi_matrix)
    print('complete_the_loading completed 2')

    content = {'matrix': matrix, 'noun_rows': noun_rows, 'verb_columns': verb_columns,
               'soc_pmi_matrix': soc_pmi_matrix, 'soc_pmi_noun_rows': soc_pmi_noun_rows,
               'soc_pmi_verb_columns': soc_pmi_verb_columns}

    return content
