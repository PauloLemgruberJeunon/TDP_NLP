"""
This file intends to be an utility box, containing functions to help with smaller functionalities
"""

import projectFiles.constants as cts

import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string

import numpy as np
import math

import os

import matplotlib.patches as mpatches
from matplotlib import pyplot as plt

import nltk
from nltk.tag.stanford import CoreNLPPOSTagger

from py4j.java_gateway import JavaGateway


'''
Below are functions called by the main.py file and are the start of the noun extraction process from books and xlsx
sheets
'''


def read_text_input(file_input_path, encoding, lowerText=False):
    print("read_text_input started")

    # This will open the text .txt archive and read it
    file = open(file_input_path, 'r', encoding=encoding)

    # This will transform the archive into a huge string
    raw_text = file.read()

    if lowerText:
        # This will lower all the upper case letters in the string (in the entire input file text)
        raw_text = raw_text.lower()

    print("read_text_input ended")

    return raw_text


def tokenize_string(text_string, eliminateNonAlphaNumericalCharacters=False):
    print("tokenize_string started")

    if eliminateNonAlphaNumericalCharacters:
        # This will create a python list named 'tokens' that will have each word/number as an element
        # It will eliminate all kinds of non-alpha numerical characters (punctuation included)
        tokens = nltk.tokenize.RegexpTokenizer(r'\w+').tokenize(text_string)
    else:
        tokens = nltk.tokenize.word_tokenize(text_string)

    print("tokenize_string ended")

    return tokens


def tag_tokens_using_stanford_corenlp(token_list, corenlp_server_address='http://localhost:9000'):
    print("tag_tokens_using_stanford_corenlp started")

    tagger = CoreNLPPOSTagger(url=corenlp_server_address)

    # The piece of code below is exists to deal with a limitation of the Stanford's coreNLP Server that only
    # supports 100000 characters per server call. So this will break the text in a lot of smaller pieces and send
    # them to the server and after will unite them all in one list of tagged words ('tagged_text')
    tagged_text = []
    txt_size = len(token_list)
    i = 0
    while i < txt_size:
        tokens_to_tag = token_list[i:i + 6000]
        i += 6001
        if i + 6000 >= txt_size:
            tokens_to_tag = token_list[i:txt_size]
            i = txt_size + 1

        tagged_text += tagger.tag(tokens_to_tag)

    print("tag_tokens_using_stanford_corenlp ended")

    return tagged_text


def tokens_to_windows(tokens, window_size):
    """
    Transform a tokenized text into a series of windows for them to serve as contexts
    :param tokens: A python list containing the tokenized text
    :param window_size: The wanted window size
    :return: A python list containing the "tokens" separated in windows
    """

    print("tokens_to_windows started")

    i = 0
    tagged_text_size = len(tokens)
    temp_list = []
    windows = []
    while i < tagged_text_size:
        temp_list.append(tokens[i])
        if (i + 1) % window_size == 0:
            windows.append(temp_list.copy())
            temp_list.clear()
        i += 1

    if tagged_text_size % window_size != 0:
        windows.append(temp_list.copy())

    print("tokens_to_windows ended")

    return windows


def tokens_to_centralized_windows(tagged_text, window_size, enable_verb_filter):
    """
    This function is other windowing method. It will find the desired verbs and make windows around them. The windows
    can overlap, its normal.
    :param tagged_text: A list of tuples. The contents of the tuples are two string, the first being the word and the
    second being the tag related to the word
    :param window_size: It is an integer that indicates the wanted size of the windows
    :return: Returns a list of lists. Each inner list is a window containing the word-tag tuples
    """

    print("tokens_to_centralized_windows started")

    tagged_text_size = len(tagged_text)
    windows = []

    lemmatizer = nltk.stem.WordNetLemmatizer()

    # Checks if the 'window_size' value is even or odd
    is_even = False
    if window_size % 2 == 0:
        is_even = True

    # Based of the window size being even or odd the code below calculates the right and left offsets from the central
    # element of the window, that is the wanted verbs in this case
    begin_offset = window_size // 2
    if is_even:
        end_offset = (window_size // 2) - 1
    else:
        end_offset = (window_size // 2)

    i = 0
    # Iterates over the tagged_text and create the windows based on it
    while i < tagged_text_size:

        # Checks if the current word is a verb and checks if it's infinitive form is present on the dict of wanted verbs
        if tagged_text[i][1].startswith('V') and (not enable_verb_filter or
           lemmatizer.lemmatize(tagged_text[i][0], 'v') in cts.verbs_to_keep):
            # Dumb but functional method to prevent windows from exceed the 'tagged-text' array's boundaries
            if i - begin_offset < 0 or i + end_offset >= tagged_text_size:
                if i - begin_offset < 0:
                    start = 0
                    end = i + end_offset
                else:
                    start = i - begin_offset
                    end = tagged_text_size - 1
            # The else holds the code for a regular execution
            else:
                start = i - begin_offset
                end = i + end_offset

            # Adds a window to the windows variables
            temp_list = [tagged_text[start:end], i-start]
            windows.append(temp_list.copy())

        i += 1

    print("tokens_to_centralized_windows ended")

    return windows

'''
Below are functions that deal with the reading, opening and saving of xlsx archives
'''

def create_workbook(name):
    #  Creates and return a Workbook with a chosen "name"
    return xlsxwriter.Workbook(name)


def get_new_worksheet(name, workbook):
    #  Add a new worksheet to a "workbook" and return it (returns the worksheet)
    return workbook.add_worksheet(name)


def close_workbook(workbook):
    #  Closes the workbook (mandatory by the end of its use)
    workbook.close()


def write_cooc_matrix(row_index_name_dict, column_index_name_dict, cooc_matrix, worksheet):
    """
    This function will print the matrix in a excel spreadsheet
    :param row_index_name_dict: The dictionary that stores the name of the rows (nouns) to print them in the sheet
    :param column_index_name_dict: The dictionary that stores the name of the columns (verbs) to print them in the sheet
    :param cooc_matrix: The co-occurrence matrix that will be written to the arquive
    :param worksheet: The worksheet that the matrix will be written to
    :return: Nothing
    """

    row_len = len(row_index_name_dict)
    column_len = len(column_index_name_dict)

    for i in range(row_len):
        for j in range(column_len):
            if i == 0:
                worksheet.write(i, j + 1, column_index_name_dict[j])
            elif j == 0:
                worksheet.write(i, j, row_index_name_dict[i])
            else:
                worksheet.write(i, j, cooc_matrix[i][j])


def write_verb_filtered_arrays(nouns_from_verb_arrays, verb_filtered_arrays, worksheet, workbook):


    xlsx_column = 0

    verbs = verb_filtered_arrays.keys()

    format = workbook.add_format({'bg_color': 'green'})

    for verb in verbs:

        nouns = nouns_from_verb_arrays[verb]
        nouns_lenght = len(nouns)

        xlsx_row = 1

        for noun in nouns:
            worksheet.write(xlsx_row, xlsx_column, noun)
            xlsx_row += 1

        xlsx_row = 0
        xlsx_column += 1
        worksheet.write(xlsx_row, xlsx_column, verb)
        xlsx_row = 1

        values = verb_filtered_arrays[verb]

        for value in values:
            worksheet.write(xlsx_row, xlsx_column, value)
            xlsx_row += 1

        xlsx_row = 0
        xlsx_column += 2

        for i in range(nouns_lenght):
            worksheet.write_blank(xlsx_row, xlsx_column, '', format)
            xlsx_row += 1

        xlsx_column += 2


def get_column_names(rows, verb_columns):
    counter = 0
    for row in rows:
        for cell in row:
            value = cell.value
            if value is not None:
                verb_columns[value] = counter
                counter += 1
        break


def complete_the_loading(rows, noun_rows, matrix):
    counter = 0
    i = 0
    for row in rows:
        row_jump = True
        j = 0

        for cell in row:
            value = cell.value
            if row_jump:
                noun_rows[value] = counter
                counter += 1
                row_jump = False
            else:
                matrix[i][j] = value
                j += 1

        i += 1


def read_all_stages():
    content_dict = {}

    wb = pd.ExcelFile(cts.path_to_interview_xlsx+ 'All_nouns_for_hypernyms_Copia.xlsx')

    wb2 = pd.ExcelFile(cts.path_to_interview_xlsx + 'list_of_verbs.xlsx')
    sheet2 = wb2.parse('sheet2')

    for i in range(1,7):
        curr_sheet = wb.parse("stage " + str(i))
        noun_list = curr_sheet['Active nouns'].values.tolist().copy()
        department_list = curr_sheet['Department'].values.tolist().copy()
        full_noun_and_verb_list = curr_sheet['Entities'].values.tolist().copy()
        synset_list = curr_sheet['SUMO word ID'].values.tolist().copy()
        final_noun_list = curr_sheet['Curated nouns'].values.tolist().copy()

        for l in range(len(noun_list)):
            if isinstance(final_noun_list[l], str):
                noun_list[l] = final_noun_list[l]

        noun_list = [str(noun_list[i]).strip() for i in range(len(noun_list))]
        department_list = [str(department_list[i]).strip() for i in range(len(department_list))]
        full_noun_and_verb_list = \
            [str(full_noun_and_verb_list[i]).strip() for i in range(len(full_noun_and_verb_list))]

        j = 0
        loop_size = len(noun_list)

        while j < loop_size:
            if isinstance(synset_list[j], float):
                if math.isnan(synset_list[j]):
                    synset_list[j] = 0.0

            if isinstance(noun_list[j], str) == False or noun_list[j] == 'nan':
                del noun_list[j:]
                del department_list[j:]
                del full_noun_and_verb_list[j:]
                del synset_list[j:]
                break
            j += 1

        synset_list = [int(k) for k in synset_list]

        # print('STAGE ' + str(i))
        # print('noun_list_length = ' + str(len(noun_list)))
        # print(noun_list)
        # print('')
        #
        # print('department_list_length = ' + str(len(department_list)))
        # print(department_list)
        # print('')
        #
        # print('full_noun_and_verb_list_length = ' + str(len(full_noun_and_verb_list)))
        # print(full_noun_and_verb_list)
        # print('')
        #
        # print('synset_list_length = ' + str(len(synset_list)))
        # print(synset_list)
        # print('\n\n')

        find_associated_verbs_in_xlsx_sheet(full_noun_and_verb_list, sheet2)

        content_dict['stage' + str(i)] = {'noun_list': noun_list, 'department_list': department_list,
                                          'synset_list': synset_list,
                                          'full_noun_and_verb_list': full_noun_and_verb_list}

    return content_dict


def read_all_nouns():
    wb = pd.ExcelFile(cts.path_to_interview_xlsx + 'All_nouns_for_hypernyms_Copia.xlsx')
    sheet = wb.parse("all nouns")
    noun_list = sheet['Active nouns'].values.tolist().copy()
    department_list = sheet['Department'].values.tolist().copy()
    full_noun_and_verb_list = sheet['Entities'].values.tolist().copy()
    synset_list = sheet['SUMO word ID'].values.tolist().copy()
    final_noun_list = sheet['Final noun'].values.tolist().copy()
    nature_of_entities_list = find_nature_of_entities(full_noun_and_verb_list, wb)

    noun_list = [str(noun_list[i]).strip() for i in range(len(noun_list))]
    department_list = [str(department_list[i]).strip() for i in range(len(department_list))]
    full_noun_and_verb_list = \
        [str(full_noun_and_verb_list[i]).strip() for i in range(len(full_noun_and_verb_list))]
    nature_of_entities_list = [str(nature_of_entities_list[i]).strip() for i in range(len(nature_of_entities_list))]

    wb3 = pd.ExcelFile(cts.path_to_interview_xlsx + 'list_of_verbs.xlsx')
    sheet3 = wb3.parse('sheet2')

    find_associated_verbs_in_xlsx_sheet(full_noun_and_verb_list, sheet3)

    lemmatizer = nltk.stem.WordNetLemmatizer()

    for i in range(len(noun_list)):
        if isinstance(final_noun_list[i], str):
            noun_list[i] = final_noun_list[i]

        noun_list[i] = noun_list[i].lower()
        noun_list[i] = lemmatizer.lemmatize(noun_list[i])

    i = 0
    whileSize = len(noun_list)
    while i < whileSize:
        noun_list[i] = noun_list[i].replace(" ", "")

        if isinstance(synset_list[i], float):
            if math.isnan(synset_list[i]):
                synset_list[i] = 0.0

        i += 1

    synset_list = [int(k) for k in synset_list]

    print(full_noun_and_verb_list)
    print(noun_list)
    print(nature_of_entities_list)
    print(department_list)
    print(synset_list)

    content_dict = {"full_noun_and_verb_list": full_noun_and_verb_list, "noun_list": noun_list,
                    "department_list": department_list, "synset_list": synset_list,
                    'nature_of_entities_list': nature_of_entities_list}

    return content_dict


def find_associated_verbs_in_xlsx_sheet(full_noun_and_verb_list, sheet):
    temp_noun_list = []
    temp_verb_list = []

    for i in range(1, 13):
        temp_noun_list += sheet['Nouns' + str(i)].values.tolist().copy()
        temp_verb_list += sheet['Verbs' + str(i)].values.tolist().copy()

    temp_noun_list = [str(temp_noun_list[i]).strip() for i in range(len(temp_noun_list))]
    temp_verb_list = [str(temp_verb_list[i]).strip() for i in range(len(temp_verb_list))]

    for j in range(len(full_noun_and_verb_list)):
        for k in range(len(temp_noun_list)):
            if full_noun_and_verb_list[j] == temp_noun_list[k]:
                full_noun_and_verb_list[j] = (full_noun_and_verb_list[j], '[' + temp_verb_list[j] + ']')
                break

    # for i in range(len(full_noun_and_verb_list)):
    #     print(full_noun_and_verb_list[i])


def find_nature_of_entities(full_noun_list, wb):
    nature_of_entities_list = []
    list_of_entities_lists = []
    list_of_nature_of_entities_list = []

    for i in range(1,7):
        curr_sheet = wb.parse("stage " + str(i))
        list_of_entities_lists.append(curr_sheet['Entities'].values.tolist().copy())
        list_of_nature_of_entities_list.append(curr_sheet['Nature of Entities (Basic)'].values.tolist().copy())

    for entity in full_noun_list:
        for i in range(6):
            should_break = False
            for j in range(len(list_of_entities_lists[i])):
                if entity == list_of_entities_lists[i][j]:
                    nature_of_entities_list.append(list_of_nature_of_entities_list[i][j].lower())
                    should_break = True
                    break

            if should_break:
                break

    return nature_of_entities_list



def load_from_wb(workbook_name):
    """
    Function used to load the matrix data from a xlsx archive. It is faster then generating it from text
    :param workbook_name: path and name of the xlsx file to use to extract information
    :return: Returns a dictionary containing the information to create a 'CoocMatrix' object
    """
    # Load workbook
    wb = load_workbook(filename=workbook_name, read_only=True)
    print('wb loaded')

    # Load worksheets
    ws = wb['cooc_matrix_full']
    ws_soc_pmi = wb['soc_pmi_matrix']

    # Calculate the matrix dimensions and transform the excel's coordinates to matrix coordinates (to only integers)
    rows = ws.rows
    matrix_dim = ws.calculate_dimension().split(':')
    rows_count = coordinate_from_string(matrix_dim[1])[1] - 1
    column_count = column_index_from_string(coordinate_from_string(matrix_dim[1])[0]) - 1

    matrix = np.empty((rows_count, column_count))

    print('matrix allocated')

    noun_rows = {}
    verb_columns = {}

    get_column_names(rows, verb_columns)
    print('get_column_names completed')
    complete_the_loading(rows, noun_rows, matrix)
    print('complete_the_loading completed')

    rows = ws_soc_pmi.rows
    matrix_dim = ws_soc_pmi.calculate_dimension().split(':')
    rows_count = coordinate_from_string(matrix_dim[1])[1] - 1
    column_count = column_index_from_string(coordinate_from_string(matrix_dim[1])[0]) - 1

    soc_pmi_matrix = np.empty((rows_count, column_count))

    soc_pmi_noun_rows = {}
    soc_pmi_verb_columns = {}

    get_column_names(rows, soc_pmi_verb_columns)
    print('get_column_names completed 2')
    complete_the_loading(rows, soc_pmi_noun_rows, soc_pmi_matrix)
    print('complete_the_loading completed 2')

    content = {'matrix': matrix, 'noun_rows': noun_rows, 'verb_columns': verb_columns,
               'soc_pmi_matrix': soc_pmi_matrix, 'soc_pmi_noun_rows': soc_pmi_noun_rows,
               'soc_pmi_verb_columns': soc_pmi_verb_columns}

    return content


'''
Below are functions related to graphics
'''


def plot_vectors(vec1_coord, vec2_coord, vec1_name, vec2_name, verb1_name, verb2_name):
    # Pyplot method to clear the old drawings
    plt.gcf().clear()

    # Calculate the module of the arrays
    vec1_module = np.sqrt(np.power(vec1_coord[0], 2) + (np.power(vec1_coord[1], 2)))
    vec2_module = np.sqrt(np.power(vec2_coord[0], 2) + (np.power(vec2_coord[1], 2)))

    # Pyplot quiver plot to show the word vectors
    plt.quiver([0, 0], [0, 0], [vec1_coord[0]/vec1_module, vec2_coord[0]/vec2_module],
               [vec1_coord[1]/vec1_module, vec2_coord[1]/vec2_module], color=['r', 'g'], angles='xy',
               scale_units='xy', scale=1)

    # Calculates the axis limits (x_begin, x_end, y_begin, y_end)
    plt.axis([0, 1, 0, 1])

    # The plotted vector's legends
    lgd_red = mpatches.Patch(color='red', label=vec1_name)
    lgd_green = mpatches.Patch(color='green', label=vec2_name)

    # Adding the legends to the plot handler
    plt.legend(handles=[lgd_red, lgd_green])

    # Adding labels to the axis
    plt.xlabel(verb1_name)
    plt.ylabel(verb2_name)

    # Shows the plot in a window but does not blocks this thread (the 'False' parameter is for the blocking option)
    plt.show(False)


'''
Below are functions related to saving in text archives
'''


def save_tagged_words(tagged_text, file_name=cts.path_to_txtFolder+'tagged_text.txt', encoding='utf8'):
    """
    Saves a list of tuples in a file. Each tuple contains two strings, the first for the word and the second for
    it's tag. This function is used to generate a file in the txtFiles folder. Use this file to check the POSTagger
    tagging accuracy
    :param tagged_text: list of tuples explained above
    :param file_name: file name to save on
    :param encoding: type of encoding of the file
    :return: None
    """
    f2 = open(file_name, 'w', encoding=encoding)
    for (word, tag) in tagged_text:
        f2.write(word + ' -> ' + tag + '\n')
    f2.close()


def save_list_of_tuples(list_of_tuples, source_name, purpose_name, encoding='utf8'):

    output_file = open(cts.path_to_txtFolder + source_name + '_' + purpose_name + '.txt', 'w', encoding=encoding)
    for tuple in list_of_tuples:
        for value in tuple:
            output_file.write(str(value) + ' - ')
        output_file.write('\n')


'''
Below are functions related to create gdf archives 
'''


def save_noun_sim_matrix_in_gdf(cooc_matrix, noun_dict, methods, book_name):

    output_files = []
    for method in methods:
        output_files.append(open(cts.path_to_book_sim_gdf + book_name + '_' + method + '.gdf', 'w'))

    for output_file_index in range(len(output_files)):
        # print('Creating the graph archives ... current progress = ' + str(output_file_index) + '/' +
        # str(len(output_files)))

        output_files[output_file_index].write('nodedef>name VARCHAR')

        for i in range(17):
            output_files[output_file_index].write(', verb' + str(i) + ' VARCHAR')

        output_files[output_file_index].write('\n')

        for noun_key in noun_dict:
            output_files[output_file_index].write(noun_key)
            still_have_verbs_to_fill = 16
            for verb_key, verb_column in cooc_matrix.verb_columns.items():
                curr_row = cooc_matrix.noun_rows[noun_key]
                if cooc_matrix.matrix[curr_row][verb_column] > 0 and still_have_verbs_to_fill > 0:
                    output_files[output_file_index].write(', ' + verb_key)
                    still_have_verbs_to_fill -= 1

            for i in range(still_have_verbs_to_fill):
                output_files[output_file_index].write(', ')

            output_files[output_file_index].write('\n')

        output_files[output_file_index].write('edgedef>node1 VARCHAR, node2 VARCHAR, weight FLOAT\n')

        inverted_noun_dict = invert_dictionary(noun_dict)

        curr_matrix = cooc_matrix.noun_to_noun_sim_matrices[methods[output_file_index]]

        i = 0
        while i < curr_matrix.shape[0] - 1:

            j = i + 1
            while j < curr_matrix.shape[0]:

                output_files[output_file_index].write(inverted_noun_dict[i] + ',' + inverted_noun_dict[j] + ',' +
                                                      '{0:.2f}'.format(curr_matrix[i][j]) + '\n')
                j += 1

            # print('archive: ' + str(i))
            i += 1


def save_noun_sim_matrix_in_gdf_2(noun_to_noun_sim_matrices, noun_list, department_list, full_noun_and_verb_list,
                                  synset_list, methods, path, name, eliminate_same_department_edges=True):
    # if eliminate_same_department_edges:
    #     dept_edges = "0"
    # else:
    #     dept_edges = '1'

    output_files = []
    for method in methods:
        output_files.append(open(path + name + '_' + method + '.gdf', 'w'))
        # output_files.append(open(path + name + '_' + method + '_' + dept_edges + '.gdf', 'w'))

    for output_file_index in range(len(output_files)):
        # print('Creating the graph archives ... current progress = ' + str(output_file_index+1) + '/' +
        #       str(len(output_files)))

        output_files[output_file_index].write('nodedef>name VARCHAR, reducedName VARCHAR, synset VARCHAR,'
                                              ' verb VARCHAR, color VARCHAR\n')

        for i in range(len(noun_list)):
            output_files[output_file_index].write(full_noun_and_verb_list[i][0] + ',' + noun_list[i] + ',' +
                                                  str(synset_list[i]) + ',' + full_noun_and_verb_list[i][1])
            department = department_list[i]

            if department in cts.department_colors:
                color = cts.department_colors[department]
            else:
                color = "#000000"

            output_files[output_file_index].write("," + color + "\n")

        output_files[output_file_index].write('edgedef>node1 VARCHAR, node2 VARCHAR, weight FLOAT\n')

        curr_matrix = noun_to_noun_sim_matrices[methods[output_file_index]]

        print("matrixShape = " + str(curr_matrix.shape[0]))

        i = 0
        while i < curr_matrix.shape[0] - 1:

            j = i + 1
            while j < curr_matrix.shape[0]:

                if eliminate_same_department_edges and department_list[i] == department_list[j]:
                    j += 1
                    continue

                output_files[output_file_index].write(full_noun_and_verb_list[i][0] + ',' +
                                                      full_noun_and_verb_list[j][0] + ',' +
                                                      '{0:.2f}'.format(curr_matrix[i][j]) + '\n')
                j += 1

            # print('archive: ' + str(i))
            i += 1


'''
Below are general use functions
'''


def invert_dictionary(dictionary):
    return dict(zip(dictionary.values(), dictionary.keys()))


def limit_value(value, min_value, max_value):
    if value > max_value:
        value = max_value
    elif value < min_value:
        value = min_value

    return value


def create_new_directory(path_plus_dir_name):
    if not os.path.exists(path_plus_dir_name):
        os.makedirs(path_plus_dir_name, exist_ok=True)


'''
Below are Java interface functions
'''


jvm_started = False


def execute_java(noun_list, department_list, full_noun_and_verb_list, synset_list, sheet_name):

    gateway = JavaGateway()
    word_container_hashmap = gateway.jvm.java.util.HashMap()

    for i in range(len(noun_list)):
        word_countainer = gateway.jvm.WordContainer()
        word_countainer.setFullWord(full_noun_and_verb_list[i][0])
        word_countainer.setVerb(full_noun_and_verb_list[i][1])
        word_countainer.setReducedWord(noun_list[i])
        word_countainer.setSynset(str(synset_list[i]).strip())
        curr_dept = department_list[i]
        if curr_dept in cts.department_colors:
            color = cts.department_colors[curr_dept]
        else:
            color = "#000000"

        word_countainer.setHexColor(color)

        codedWord = ''
        for letter in word_countainer.getFullWord():
            codedWord += str(ord(letter))

        word_container_hashmap.put(codedWord + word_countainer.getSynset(), word_countainer)

    word_graph_handler = gateway.entry_point
    word_graph = word_graph_handler.getWordGraph(word_container_hashmap,
                                                 cts.path_to_interview_hypernym_gdf + sheet_name + "_hyperGraph.gdf")

    word_graph.startWordGraph()


'''
Below here are some pre-steps that have to be made
'''


create_new_directory(cts.path_to_gdfFolder)
create_new_directory(cts.path_to_xlsxFolder)
create_new_directory(cts.path_to_txtFolder)
create_new_directory(cts.path_to_book_sim_gdf)
create_new_directory(cts.path_to_interview_hypernym_gdf)
create_new_directory(cts.path_to_interview_sim_gdf)
create_new_directory(cts.path_to_interview_xlsx)
create_new_directory(cts.path_to_generated_xlsx)
