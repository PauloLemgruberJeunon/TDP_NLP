"""
This file intends to be an utility box, containing functions to help with smaller functionalities
"""

import xlsxwriter
import numpy as np
from matplotlib import pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string


verbs_to_keep = {'prepare': 1, 'synthesize': 1, 'generate': 1, 'define': 1, 'illustrate': 1, 'classify': 1,
                'develop': 1, 'name': 1, 'defend': 1, 'explain': 1, 'describe': 1, 'criticize': 1,
                'test': 1, 'review': 1, 'order': 1, 'analyze': 1, 'choose': 1, 'create': 1, 'combine': 1, 'infer': 1,
                'extend': 1, 'modify': 1, 'compare': 1, 'indicate': 1, 'distinguish': 1, 'interpret': 1, 'justify': 1,
                'identify': 1, 'list': 1, 'evaluate': 1, 'calculate': 1, 'design': 1, 'recognize': 1, 'model': 1,
                'discuss': 1, 'practice': 1, 'apply': 1, 'estimate': 1, 'compute': 1, 'solve': 1, 'conclude': 1,
                'predict': 1}


def measure_postag_accuracy(tagged_sents, tagged_words):
    """
    Function used to measure the accuracy of a POSTagger (does not handle different types of tag standards)
    Parameters
    ----------
    tagged_sents : Are the sentences of the database used to measure the accuracy of the tagger ("Test set")
    tagged_words : Are the words of the sentences stored in "tagged_sents" the were tagged by the POSTagger

    Returns
    -------
    The measured accuracy of the POSTagger
    """

    total_count = 0
    right_count = 0

    for i in range(0, len(tagged_sents)):
        print('[Accuracy] Current stage = ' + str(i))
        k = 0
        j = 0
        local_count = 0
        local_right_count = 0
        while j < len(tagged_sents[i]):
            total_count += 1
            local_count += 1
            if tagged_sents[i][j][1] == '-NONE-':
                while tagged_sents[i][j][1] == '-NONE-':
                    j += 1
                while tagged_sents[i][j][0] != tagged_words[i][k][0]:
                    k += 1
            if tagged_sents[i][j][1] == tagged_words[i][k][1]:
                right_count += 1
                local_right_count += 1

            k += 1
            j += 1
        print('Local iter accuracy = ' + str(local_right_count/local_count) + '\n')

    return right_count/total_count


def create_workbook(name):
    #  Creates and return a Workbook with a chosen "name"
    return xlsxwriter.Workbook(name)


def get_new_worksheet(name, workbook):
    #  Add a new worksheet to a "workbook" and return it (returns the worksheet)
    return workbook.add_worksheet(name)


def close_workbook(workbook):
    #  Closes the workbook (mandatory by the end of its use)
    workbook.close()


def write_cooc_matrix(row_index_name_dict, column_index_name_dict, cooc_matrix, worksheet):
    """
    This function will print the matrix in a excel spreadsheet
    :param row_index_name_dict: The dictionary that stores the name of the rows (nouns) to print them in the sheet
    :param column_index_name_dict: The dictionary that stores the name of the columns (verbs) to print them in the sheet
    :param cooc_matrix: The co-occurrence matrix that will be written to the arquive
    :param worksheet: The worksheet that the matrix will be written to
    :return: Nothing
    """

    row_len = len(row_index_name_dict)
    column_len = len(column_index_name_dict)

    for i in range(row_len):
        worksheet.write(i + 1, 0, row_index_name_dict[i])

    for j in range(column_len):
        worksheet.write(0, j + 1, column_index_name_dict[j])

    for i in range(row_len):
        for j in range(column_len):
            worksheet.write(i + 1, j + 1, cooc_matrix[i][j])


def tokens_to_windows(tokens, window_size):
    """
    Transform a tokenized text into a series of windows for them to serve as contexts
    :param window_size: A python list containing a tokenized text
    :return: A python list containing the "tokens" separated in windows
    """
    i = 0
    tagged_text_size = len(tokens)
    temp_list = []
    windows = []
    while i < tagged_text_size:
        temp_list.append(tokens[i])
        if (i + 1) % window_size == 0:
            windows.append(temp_list.copy())
            temp_list.clear()
        i += 1

    if tagged_text_size % window_size != 0:
        windows.append(temp_list.copy())

    return windows


def tokens_to_centralized_windows(tokens, window_size):
    tagged_text_size = len(tokens)
    windows = []

    is_even = False
    if window_size % 2 == 0:
        is_even = True

    begin_offset = window_size // 2
    if is_even:
        end_offset = (window_size // 2) - 1
    else:
        end_offset = (window_size // 2)

    i = 0
    while i < tagged_text_size:
        if tokens[i][0] in verbs_to_keep:
            if i - begin_offset < 0 or i + end_offset >= tagged_text_size:
                if i - begin_offset < 0:
                    start = 0
                    end = i + end_offset
                else:
                    start = i - begin_offset
                    end = tagged_text_size - 1
            else:
                start = i - begin_offset
                end = i + end_offset

            temp_list = tokens[start:end]
            windows.append(temp_list.copy())

        i += 1

    return windows


def invert_dictionary(dictionary):
    return dict(zip(dictionary.values(), dictionary.keys()))


def plot_vectors(vec1_coord, vec2_coord, vec1_name, vec2_name, verb1_name, verb2_name):
    plt.gcf().clear()
    vec1_module = np.sqrt(np.power(vec1_coord[0], 2) + np.power(vec1_coord[1], 2))
    vec2_module = np.sqrt(np.power(vec2_coord[0], 2) + (np.power(vec2_coord[1], 2)))
    plt.quiver([0, 0], [0, 0], [vec1_coord[0]/vec1_module, vec2_coord[0]/vec2_module],
               [vec1_coord[1]/vec1_module, vec2_coord[1]/vec2_module], color=['r', 'g'], angles='xy',
               scale_units='xy', scale=1)
    plt.axis([0, 1, 0, 1])
    lgd_red = mpatches.Patch(color='red', label=vec1_name)
    lgd_green = mpatches.Patch(color='green', label=vec2_name)
    plt.legend(handles=[lgd_red, lgd_green])
    plt.xlabel(verb1_name)
    plt.ylabel(verb2_name)
    plt.show(False)


def get_column_names(rows, verb_columns):

    counter = 0
    for row in rows:
        for cell in row:
            value = cell.value
            if value is not None:
                verb_columns[value] = counter
                counter += 1
        break


def complete_the_loading(rows, noun_rows, matrix):
    counter = 0
    i = 0
    for row in rows:
        row_jump = True
        j = 0

        for cell in row:
            value = cell.value
            if row_jump:
                noun_rows[value] = counter
                counter += 1
                row_jump = False
            else:
                matrix[i][j] = value
                j += 1

        i += 1


def load_from_wb(workbook_name):
    wb = load_workbook(filename=workbook_name, read_only=True)
    ws = wb['cooc_matrix_full']
    ws_filtered = wb['cooc_matrix_filtered']

    rows = ws.rows
    matrix_dim = ws.calculate_dimension().split(':')
    rows_count = coordinate_from_string(matrix_dim[1])[1] - 1
    column_count = column_index_from_string(coordinate_from_string(matrix_dim[1])[0]) - 1

    matrix = np.empty((rows_count, column_count))

    noun_rows = {}
    verb_columns = {}

    get_column_names(rows, verb_columns)
    complete_the_loading(rows, noun_rows, matrix)

    rows = ws_filtered.rows
    matrix_dim = ws_filtered.calculate_dimension().split(':')
    rows_count = coordinate_from_string(matrix_dim[1])[1] - 1
    column_count = column_index_from_string(coordinate_from_string(matrix_dim[1])[0]) - 1

    filtered_matrix = np.empty((rows_count, column_count))

    filtered_noun_rows = {}
    filtered_verb_columns = {}

    get_column_names(rows, filtered_verb_columns)
    complete_the_loading(rows, filtered_noun_rows, filtered_matrix)

    content = {'matrix': matrix, 'noun_rows': noun_rows, 'verb_columns': verb_columns,
               'filtered_matrix': filtered_matrix, 'filtered_noun_rows': filtered_noun_rows,
               'filtered_verb_columns': filtered_verb_columns}

    return content


def save_tagged_words(tagged_text, file_name='tagged_text.txt', encoding='utf8'):
    f2 = open(file_name, 'w', encoding=encoding)
    for (word, tag) in tagged_text:
        f2.write(word + ' -> ' + tag + '\n')
    f2.close()


def pause():
    input('Press enter to continue')
